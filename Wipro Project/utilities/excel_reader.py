# import os
#
# from openpyxl import load_workbook
#
#
# def get_test_data():
#     current_dir = os.path.dirname(__file__)
#
#     file_path = os.path.join(
#         current_dir,
#         "..",
#         # "test_data",
#         "search_data.xlsx"
#     )
#
#     workbook = load_workbook(file_path)
#     sheet = workbook.active
#
#     data = {
#         "mobile_number": sheet["A2"].value,
#         "location": sheet["B2"].value
#     }
#
#     return data



import os
from openpyxl import load_workbook

def get_test_data():
    file_path = os.path.join(
        os.path.dirname(__file__),
        "..",
        # "test_data",
        "search_data.xlsx"
    )

    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            "location": row[0],
            "property_type": row[1],
            "budget": row[2]
        })

    return data
