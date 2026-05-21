from __future__ import annotations

import copy
import json
import os
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from config.settings import ROOT_DIR
from utilities.excel_reader import get_test_data


DEFAULT_TEST_DATA_PATH = ROOT_DIR / "test_data" / "test_data.xml"
EXCEL_XML_NAMESPACE = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}


class TestDataLoader:
    @staticmethod
    def load(path: str | Path | None = None) -> dict[str, Any]:
        data_path = Path(path) if path else DEFAULT_TEST_DATA_PATH
        if not data_path.is_absolute():
            data_path = ROOT_DIR / data_path

        if not data_path.exists():
            raise FileNotFoundError(f"Test data file was not found: {data_path}")

        suffix = data_path.suffix.lower()
        if suffix == ".json":
            data = TestDataLoader._load_json(data_path)
        elif suffix == ".xlsx":
            data = TestDataLoader._load_xlsx(data_path)
        elif suffix == ".xml":
            data = TestDataLoader._load_excel_xml(data_path)
        else:
            raise ValueError(
                f"Unsupported test data file format: {data_path}. Use Excel (.xlsx), Excel XML (.xml), or JSON (.json)."
            )

        if not isinstance(data, dict):
            raise ValueError(f"Test data file must contain a top-level object: {data_path}")

        return TestDataLoader._with_secret_overrides(data)

    @staticmethod
    def _load_json(data_path: Path) -> dict[str, Any]:
        with data_path.open(encoding="utf-8") as file:
            return json.load(file)

    @staticmethod
    def _load_xlsx(data_path: Path) -> dict[str, Any]:
        workbook = load_workbook(data_path)
        if workbook.worksheets:
            worksheet = workbook.active
            headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in worksheet[1]]
            if headers[:3] == ["key", "value", "type"]:
                data: dict[str, Any] = {}
                for sheet in workbook.worksheets:
                    rows = list(sheet.iter_rows(values_only=True))
                    if not rows:
                        continue

                    section_headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
                    if section_headers[:3] != ["key", "value", "type"]:
                        continue

                    section_data: dict[str, Any] = {}
                    for row in rows[1:]:
                        normalized_row = list(row) + [None] * (3 - len(row))
                        key = str(normalized_row[0]).strip() if normalized_row[0] is not None else ""
                        if not key:
                            continue
                        raw_value = normalized_row[1]
                        raw_type = str(normalized_row[2]).strip().lower() if normalized_row[2] is not None else "str"
                        section_data[key] = TestDataLoader._coerce_value(raw_value, raw_type or "str", sheet.title, key)
                    data[sheet.title] = section_data

                if data:
                    return data

        excel_data = get_test_data()
        return {
            "login": {
                "mobile_number": excel_data["mobile_number"],
                "username": "",
                "password": "",
                "before_mobile_entry_pause_seconds": 0,
                "after_mobile_entry_pause_seconds": 0,
                "otp_pause_seconds": 60,
            },
            "property_search": {
                "default_location": excel_data["location"],
                "filtered_location": excel_data["location"],
            },
            "navigation": {
                "primary_menu_items": [],
            },
        }

    @staticmethod
    def _load_excel_xml(data_path: Path) -> dict[str, Any]:
        workbook = ET.parse(data_path)
        root = workbook.getroot()
        data: dict[str, Any] = {}

        for worksheet in root.findall(".//ss:Worksheet", EXCEL_XML_NAMESPACE):
            sheet_name = worksheet.attrib.get(f"{{{EXCEL_XML_NAMESPACE['ss']}}}Name")
            if not sheet_name:
                continue

            rows = TestDataLoader._worksheet_rows(worksheet)
            if not rows:
                continue

            headers = [str(value).strip().lower() for value in rows[0]]
            if headers[:3] != ["key", "value", "type"]:
                raise ValueError(
                    f"Worksheet '{sheet_name}' must start with headers: key, value, type. Found: {headers}"
                )

            sheet_data: dict[str, Any] = {}
            for row in rows[1:]:
                normalized_row = row + [""] * (3 - len(row))
                key = str(normalized_row[0]).strip()
                if not key:
                    continue

                raw_value = normalized_row[1]
                raw_type = str(normalized_row[2]).strip().lower() or "str"
                sheet_data[key] = TestDataLoader._coerce_value(raw_value, raw_type, sheet_name, key)

            data[sheet_name] = sheet_data

        return data

    @staticmethod
    def _worksheet_rows(worksheet: ET.Element) -> list[list[str]]:
        rows: list[list[str]] = []
        for row in worksheet.findall(".//ss:Table/ss:Row", EXCEL_XML_NAMESPACE):
            row_values: list[str] = []
            expected_index = 1

            for cell in row.findall("ss:Cell", EXCEL_XML_NAMESPACE):
                index_value = cell.attrib.get(f"{{{EXCEL_XML_NAMESPACE['ss']}}}Index")
                if index_value:
                    target_index = int(index_value)
                    while expected_index < target_index:
                        row_values.append("")
                        expected_index += 1

                data_element = cell.find("ss:Data", EXCEL_XML_NAMESPACE)
                row_values.append("" if data_element is None or data_element.text is None else data_element.text)
                expected_index += 1

            if any(str(value).strip() for value in row_values):
                rows.append(row_values)

        return rows

    @staticmethod
    def _coerce_value(raw_value: Any, raw_type: str, sheet_name: str, key: str) -> Any:
        text_value = "" if raw_value is None else str(raw_value).strip()

        if raw_type in {"str", "string", "text"}:
            return text_value
        if raw_type in {"int", "integer"}:
            return int(text_value) if text_value else 0
        if raw_type in {"float", "decimal", "number"}:
            return float(text_value) if text_value else 0.0
        if raw_type in {"bool", "boolean"}:
            return text_value.lower() in {"1", "true", "yes", "y", "on"}
        if raw_type == "json":
            return json.loads(text_value) if text_value else {}
        if raw_type == "list":
            if not text_value:
                return []
            try:
                parsed = json.loads(text_value)
            except json.JSONDecodeError as exc:
                raise ValueError(
                    f"Worksheet '{sheet_name}' key '{key}' declared as list but does not contain valid JSON."
                ) from exc
            if not isinstance(parsed, list):
                raise ValueError(f"Worksheet '{sheet_name}' key '{key}' must contain a JSON array for type=list.")
            return parsed

        raise ValueError(f"Unsupported type '{raw_type}' for worksheet '{sheet_name}' key '{key}'.")

    @staticmethod
    def _with_secret_overrides(data: dict[str, Any]) -> dict[str, Any]:
        merged_data = copy.deepcopy(data)
        login_data = merged_data.setdefault("login", {})

        username = os.getenv("ACRES_USERNAME")
        password = os.getenv("ACRES_PASSWORD")
        if username:
            login_data["username"] = username
            login_data["mobile_number"] = username
        if password:
            login_data["password"] = password

        return merged_data
